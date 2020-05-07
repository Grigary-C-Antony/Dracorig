import os
import time
import re
from urllib.request import urlopen, Request
from datetime import datetime
from openpyxl import Workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
g=[]
print("You need a chrome browser")
time.sleep(1)
q1 = str(input("type website :"))
q2 = str(input("type proffession :"))
q3 = str(input("type country:"))
q4 = str(input("type email extension :"))
q5 = str(input("type alt email extension :"))
time.sleep(1)
print("pls Don't close the chrome browser")
print("pls Don't close the chrome browser")
print("pls Don't close the chrome browser")
time.sleep(2)


M = 'sites:'+q1+' "'+q2+'" "'+q3+'" "'+q4+'" , "'+q5+'"'
driver = webdriver.Chrome("chromedriver.exe")
driver.maximize_window()
# driver.find_element_by_xpath("/html/body").send_keys(Keys.F11)
driver.set_page_load_timeout(50)
driver.get("http://www.google.com")
ele = driver.find_element_by_name("q")  # search bar
ele.send_keys(M)
time.sleep(2)
ele.send_keys(Keys.ENTER)
driver.refresh()
print(driver.current_url)
for i in range(3,10):
    k=str(i)
    # print(k)
    driver.find_element_by_xpath("//*[@id='xjs']/div/table/tbody/tr/td["+ k +"]/a").click()
    # driver.find
    # print(driver.current_url)
    g.append(driver.current_url)
    time.sleep(5)
driver.quit()

save_excel = True  # Change to "True" to save email into Excel

book = Workbook()
sheet = book.active


def start_scrape(page, name_the_file):
    print("\n\nWebpage is currently being scrapped... please wait...")

    scrape = BeautifulSoup(page, 'html.parser')
    scrape = scrape.get_text()

    phone_numbers = set(
        re.findall(r"((?:\d{3}|\(\d{3}\))?(?:\s|-|\.)?\d{3}(?:\s|-|\.)\d{4})", scrape))  # "set" removes duplicates
    emails = set(re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,3}", scrape))

    nodupnumber = len(list(phone_numbers))
    nodupemail = len(list(emails))

    dupnumber = len(list(re.findall(r"((?:\d{3}|\(\d{3}\))?(?:\s|-|\.)?\d{3}(?:\s|-|\.)\d{4})", scrape)))
    dupemail = len(list(re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,3}", scrape)))

    number_of_dup_number = int(dupnumber) - int(nodupnumber)
    number_of_dup_email = int(dupemail) - int(nodupemail)

    email_list = list(emails)

    if len(phone_numbers) == 0:
        print("No phone number(s) found.")

        print("-----------------------------\n")
    else:
        count = 1
        for item in phone_numbers:
            print("Phone number #" + str(count) + ': ' + item)
            count += 1

    print("-----------------------------\n")

    if len(emails) == 0:
        print("No email address(es) found.")
        print("-----------------------------\n")
    else:
        count = 1
        for item in emails:

            print('Email address #' + str(count) + ': ' + item.lower())
            count += 1

    if save_excel:
        for row in zip(email_list):
            sheet.append(row)
        excel_file = (name_the_file + ".xlsx")
        book.save(excel_file)

    # print("\nDuplicates have been removed from list.")
    # print("Total phone numbers: ", nodupnumber)
    # print("Total email addresses: ", nodupemail)
    # print("There were " + str(number_of_dup_number) + " duplicate phone numbers.")
    # print("There were " + str(number_of_dup_email) + " duplicate email addresses.")

    if save_excel:
        print("\n\nData has been stored inside of an Excel spreadsheet named: " + excel_file + " in this directory: " + os.getcwd())
        mod_time = os.stat(excel_file).st_mtime

        print("\nTry Completed: " + str(datetime.fromtimestamp(mod_time))+"  Going to next Try")
        # print("\nSize of file: " + str(os.stat(excel_file).st_size) + " KB")


def main():

    if save_excel:
        name_the_file = input("Name the file you would like to save the data in (don't include .xlsx): ")

    for ben in g :
        webpage = str(ben)
        try:
            page = urlopen(webpage)
            start_scrape(page)
        except:
            hdr = {'User-Agent': 'Mozilla/5.0'}
            req = Request(webpage, headers=hdr)
            page = urlopen(req)
            start_scrape(page, name_the_file)

if __name__ == "__main__":
    main()

h = []

wb = load_workbook('jjjj.xlsx')

sheet = wb.active

m_row = sheet.max_row

for i in range(1, m_row):
    cell = sheet.cell(row=i, column=1)

    h.append(cell.value)

h.sort()
print(h)
h.save(filename='data.xlsx')
